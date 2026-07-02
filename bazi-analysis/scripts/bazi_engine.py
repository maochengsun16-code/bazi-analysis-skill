# -*- coding: utf-8 -*-
from __future__ import annotations

import csv
import json
import sys
from dataclasses import dataclass
from io import BytesIO, StringIO
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional


ROOT = Path(__file__).resolve().parent
VENDOR = ROOT / "vendor"
if str(VENDOR) not in sys.path:
    sys.path.insert(0, str(VENDOR))

from lunar_python import Lunar, Solar  # noqa: E402


GAN_YANG = {"甲", "丙", "戊", "庚", "壬"}
POSITIONS = ("year", "month", "day", "time")
POSITION_LABELS = {"year": "年", "month": "月", "day": "日", "time": "时"}
TOMB_BRANCHES = {"辰", "戌", "丑", "未"}
WINTER_BRANCHES = {"亥", "子", "丑"}
SUMMER_BRANCHES = {"巳", "午", "未"}

ROOT_RULES = {
    "木": {
        "day_masters": {"甲", "乙"},
        "bijie": {"寅", "卯"},
        "yin": {"亥", "子"},
        "storehouse": {"辰", "未"},
    },
    "火": {
        "day_masters": {"丙", "丁"},
        "bijie": {"巳", "午"},
        "yin": {"寅", "卯"},
        "storehouse": {"戌", "未"},
    },
    "土": {
        "day_masters": {"戊", "己"},
        "bijie": {"辰", "戌", "丑", "未"},
        "yin": {"巳", "午"},
        "storehouse": {"辰", "戌", "丑", "未"},
        "support": {"巳", "午"},
    },
    "金": {
        "day_masters": {"庚", "辛"},
        "bijie": {"申", "酉"},
        "yin": {"辰", "戌", "丑", "未"},
        "storehouse": {"丑", "戌"},
    },
    "水": {
        "day_masters": {"壬", "癸"},
        "bijie": {"亥", "子"},
        "yin": {"申", "酉"},
        "storehouse": {"辰", "丑"},
    },
}

PALACE_LABELS = {
    "year": "年柱/祖上宫",
    "month": "月柱/父母兄弟宫",
    "day": "日柱/日主夫妻宫",
    "time": "时柱/子女晚年宫",
}

STEM_COMBINATIONS = {
    frozenset(("甲", "己")): ("甲己合", "土"),
    frozenset(("乙", "庚")): ("乙庚合", "金"),
    frozenset(("丙", "辛")): ("丙辛合", "水"),
    frozenset(("丁", "壬")): ("丁壬合", "木"),
    frozenset(("戊", "癸")): ("戊癸合", "火"),
}

STEM_CLASHES = {
    frozenset(("甲", "庚")): "甲庚冲",
    frozenset(("乙", "辛")): "乙辛冲",
    frozenset(("丙", "壬")): "丙壬冲",
    frozenset(("丁", "癸")): "丁癸冲",
}

BRANCH_LIU_HE = {
    frozenset(("子", "丑")): ("子丑合", "土", "闭丑金库"),
    frozenset(("寅", "亥")): ("寅亥合", "木", "破合，表面合作内里算计"),
    frozenset(("卯", "戌")): ("卯戌合", "火", "闭戌火库"),
    frozenset(("辰", "酉")): ("辰酉合", "金", "闭辰水库"),
    frozenset(("巳", "申")): ("巳申合", "水", "破合，表面合作内里算计"),
    frozenset(("午", "未")): ("午未合", "火", "闭未木库"),
}

BRANCH_SAN_HE = {
    frozenset(("申", "子", "辰")): ("申子辰三合水局", "水"),
    frozenset(("亥", "卯", "未")): ("亥卯未三合木局", "木"),
    frozenset(("寅", "午", "戌")): ("寅午戌三合火局", "火"),
    frozenset(("巳", "酉", "丑")): ("巳酉丑三合金局", "金"),
}

BRANCH_HALF_HE = {
    frozenset(("申", "子")): ("申子半合水", "水", "生地半合"),
    frozenset(("子", "辰")): ("子辰半合水", "水", "墓地半合"),
    frozenset(("亥", "卯")): ("亥卯半合木", "木", "生地半合"),
    frozenset(("卯", "未")): ("卯未半合木", "木", "墓地半合"),
    frozenset(("寅", "午")): ("寅午半合火", "火", "生地半合"),
    frozenset(("午", "戌")): ("午戌半合火", "火", "墓地半合"),
    frozenset(("巳", "酉")): ("巳酉半合金", "金", "生地半合"),
    frozenset(("酉", "丑")): ("酉丑半合金", "金", "墓地半合"),
}

BRANCH_SAN_HUI = {
    frozenset(("寅", "卯", "辰")): ("寅卯辰三会木", "木"),
    frozenset(("巳", "午", "未")): ("巳午未三会火", "火"),
    frozenset(("申", "酉", "戌")): ("申酉戌三会金", "金"),
    frozenset(("亥", "子", "丑")): ("亥子丑三会水", "水"),
}

BRANCH_CHONG = {
    frozenset(("子", "午")): "子午冲",
    frozenset(("丑", "未")): "丑未冲",
    frozenset(("寅", "申")): "寅申冲",
    frozenset(("卯", "酉")): "卯酉冲",
    frozenset(("辰", "戌")): "辰戌冲",
    frozenset(("巳", "亥")): "巳亥冲",
}

BRANCH_HAI_CHUAN = {
    frozenset(("寅", "巳")): "寅巳穿",
    frozenset(("子", "未")): "子未穿",
    frozenset(("卯", "辰")): "卯辰穿",
    frozenset(("酉", "戌")): "酉戌穿",
    frozenset(("申", "亥")): "申亥穿",
    frozenset(("丑", "午")): "丑午穿",
}

BRANCH_PO = {
    frozenset(("子", "酉")): "子酉破",
    frozenset(("寅", "亥")): "寅亥破",
    frozenset(("卯", "午")): "卯午破",
    frozenset(("辰", "丑")): "辰丑破",
    frozenset(("巳", "申")): "巳申破",
    frozenset(("未", "戌")): "未戌破",
}

BRANCH_JUE = {
    frozenset(("子", "巳")): "子巳绝",
    frozenset(("亥", "午")): "亥午绝",
    frozenset(("卯", "申")): "卯申绝",
    frozenset(("寅", "酉")): "寅酉绝",
}

BRANCH_THREE_PUNISHMENTS = {
    frozenset(("丑", "未", "戌")): "丑未戌三刑",
    frozenset(("寅", "巳", "申")): "寅巳申三刑",
}

BRANCH_SELF_PUNISH = {"亥", "辰", "酉", "午"}


@dataclass(frozen=True)
class BaziRequest:
    calendar_type: str
    year: int
    month: int
    day: int
    hour: int
    minute: int = 0
    second: int = 0
    gender: str = "male"
    is_leap_month: bool = False
    sect: int = 2
    yun_sect: int = 1
    da_yun_count: int = 10
    name: str = ""


def normalize_gender(value: Any) -> str:
    text = str(value or "").strip().lower()
    if text in {"男", "男性", "male", "m", "1"}:
        return "male"
    if text in {"女", "女性", "female", "f", "0"}:
        return "female"
    raise ValueError("性别只支持 男/女")


def gender_number(gender: str) -> int:
    return 1 if normalize_gender(gender) == "male" else 0


def normalize_calendar_type(value: Any) -> str:
    text = str(value or "").strip().lower()
    if text in {"阳历", "公历", "solar", "gregorian"}:
        return "solar"
    if text in {"农历", "阴历", "lunar"}:
        return "lunar"
    raise ValueError("日期类型只支持 阳历/农历")


def parse_bool(value: Any) -> bool:
    if isinstance(value, bool):
        return value
    text = str(value or "").strip().lower()
    return text in {"是", "真", "true", "1", "yes", "y", "闰", "闰月"}


def _int_value(data: Dict[str, Any], keys: Iterable[str], default: Optional[int] = None) -> int:
    for key in keys:
        value = data.get(key)
        if value is not None and str(value).strip() != "":
            return int(float(str(value).strip()))
    if default is not None:
        return default
    raise ValueError("缺少字段：" + "/".join(keys))


def request_from_dict(data: Dict[str, Any]) -> BaziRequest:
    calendar_type = normalize_calendar_type(
        data.get("calendar_type", data.get("calendarType", data.get("日期类型", data.get("历法", "阳历"))))
    )
    gender = normalize_gender(data.get("gender", data.get("性别", "男")))
    return BaziRequest(
        name=str(data.get("name", data.get("姓名", ""))).strip(),
        calendar_type=calendar_type,
        year=_int_value(data, ("year", "年", "出生年")),
        month=_int_value(data, ("month", "月", "出生月")),
        day=_int_value(data, ("day", "日", "出生日")),
        hour=_int_value(data, ("hour", "时", "小时", "出生时"), 0),
        minute=_int_value(data, ("minute", "分", "分钟", "出生分"), 0),
        second=_int_value(data, ("second", "秒"), 0),
        gender=gender,
        is_leap_month=parse_bool(data.get("is_leap_month", data.get("isLeapMonth", data.get("是否闰月", "")))),
        sect=_int_value(data, ("sect", "八字流派"), 2),
        yun_sect=_int_value(data, ("yun_sect", "yunSect", "起运算法"), 1),
        da_yun_count=_int_value(data, ("da_yun_count", "daYunCount", "大运数量"), 10),
    )


def _validate_request(req: BaziRequest) -> None:
    if req.sect not in {1, 2}:
        raise ValueError("八字流派 sect 只支持 1 或 2")
    if req.yun_sect not in {1, 2}:
        raise ValueError("起运算法 yun_sect 只支持 1 或 2")
    if not 0 <= req.hour <= 23:
        raise ValueError("小时必须在 0-23 之间")
    if not 0 <= req.minute <= 59:
        raise ValueError("分钟必须在 0-59 之间")
    if not 0 <= req.second <= 59:
        raise ValueError("秒必须在 0-59 之间")
    if req.da_yun_count < 1 or req.da_yun_count > 12:
        raise ValueError("大运数量建议在 1-12 之间")


def _solar_lunar_from_request(req: BaziRequest):
    _validate_request(req)
    if req.calendar_type == "solar":
        solar = Solar.fromYmdHms(req.year, req.month, req.day, req.hour, req.minute, req.second)
        lunar = solar.getLunar()
    else:
        lunar_month = -abs(req.month) if req.is_leap_month else req.month
        lunar = Lunar.fromYmdHms(req.year, lunar_month, req.day, req.hour, req.minute, req.second)
        solar = lunar.getSolar()
    return solar, lunar


def _pillar(eight_char, name: str) -> Dict[str, Any]:
    getters = {
        "year": ("Year", "年柱"),
        "month": ("Month", "月柱"),
        "day": ("Day", "日柱"),
        "time": ("Time", "时柱"),
    }
    suffix, label = getters[name]
    return {
        "label": label,
        "gan_zhi": getattr(eight_char, f"get{suffix}")(),
        "gan": getattr(eight_char, f"get{suffix}Gan")(),
        "zhi": getattr(eight_char, f"get{suffix}Zhi")(),
        "wu_xing": getattr(eight_char, f"get{suffix}WuXing")(),
        "na_yin": getattr(eight_char, f"get{suffix}NaYin")(),
        "shi_shen_gan": getattr(eight_char, f"get{suffix}ShiShenGan")(),
        "shi_shen_zhi": getattr(eight_char, f"get{suffix}ShiShenZhi")(),
        "hide_gan": getattr(eight_char, f"get{suffix}HideGan")(),
        "di_shi": getattr(eight_char, f"get{suffix}DiShi")(),
        "xun": getattr(eight_char, f"get{suffix}Xun")(),
        "xun_kong": getattr(eight_char, f"get{suffix}XunKong")(),
    }


def _fmt_lunar(lunar) -> str:
    leap = "闰" if lunar.getMonth() < 0 else ""
    return f"{lunar.getYearInChinese()}年{leap}{lunar.getMonthInChinese()}月{lunar.getDayInChinese()} {lunar.getTimeZhi()}时"


def _fmt_start(yun) -> str:
    return f"{yun.getStartYear()}年{yun.getStartMonth()}个月{yun.getStartDay()}天{yun.getStartHour()}小时"


def _liu_yue_list(liu_nian) -> List[Dict[str, Any]]:
    return [
        {
            "index": ly.getIndex() + 1,
            "month": ly.getMonthInChinese() + "月",
            "gan_zhi": ly.getGanZhi(),
            "xun": ly.getXun(),
            "xun_kong": ly.getXunKong(),
        }
        for ly in liu_nian.getLiuYue()
    ]


def _liu_nian_list(da_yun) -> List[Dict[str, Any]]:
    return [
        {
            "index": ln.getIndex() + 1,
            "year": ln.getYear(),
            "age": ln.getAge(),
            "gan_zhi": ln.getGanZhi(),
            "xun": ln.getXun(),
            "xun_kong": ln.getXunKong(),
            "liu_yue": _liu_yue_list(ln),
        }
        for ln in da_yun.getLiuNian()
    ]


def _position_item(pillars: Dict[str, Dict[str, Any]], position: str, part: str) -> Dict[str, str]:
    pillar = pillars[position]
    return {
        "position": position,
        "label": f"{POSITION_LABELS[position]}{part}",
        "pillar": pillar["gan_zhi"],
        "palace": PALACE_LABELS[position],
        "value": pillar["gan"] if part == "干" else pillar["zhi"],
    }


def _palace_relation(items: List[Dict[str, str]]) -> str:
    return " 与 ".join(f"{item['label']}({item['palace']})" for item in items)


def _pair_relation(
    relation_type: str,
    name: str,
    items: List[Dict[str, str]],
    **extra: Any,
) -> Dict[str, Any]:
    relation = {
        "type": relation_type,
        "name": name,
        "positions": items,
        "palace_relation": _palace_relation(items),
        "involves_day_pillar": any(item["position"] == "day" for item in items),
    }
    relation.update(extra)
    return relation


def _stem_relations(pillars: Dict[str, Dict[str, Any]]) -> Dict[str, Any]:
    stems = [(position, pillars[position]["gan"]) for position in POSITIONS]
    combines = []
    clashes = []
    for left_index, (left_position, left_stem) in enumerate(stems):
        for right_position, right_stem in stems[left_index + 1 :]:
            key = frozenset((left_stem, right_stem))
            items = [
                _position_item(pillars, left_position, "干"),
                _position_item(pillars, right_position, "干"),
            ]
            if key in STEM_COMBINATIONS:
                name, transform = STEM_COMBINATIONS[key]
                combines.append(
                    _pair_relation(
                        "天干五合",
                        name,
                        items,
                        stems=[left_stem, right_stem],
                        transform_element=transform,
                        involves_day_master=any(item["position"] == "day" for item in items),
                        note="合不等于化，是否合化需另看月令、化神根气和地支配合",
                    )
                )
            if key in STEM_CLASHES:
                clashes.append(
                    _pair_relation(
                        "天干四冲",
                        STEM_CLASHES[key],
                        items,
                        stems=[left_stem, right_stem],
                        involves_day_master=any(item["position"] == "day" for item in items),
                    )
                )
    return {"combinations": combines, "clashes": clashes}


def _branch_items_by_value(pillars: Dict[str, Dict[str, Any]], values: set[str]) -> List[Dict[str, str]]:
    return [_position_item(pillars, position, "支") for position in POSITIONS if pillars[position]["zhi"] in values]


def _branch_pair_relations(pillars: Dict[str, Dict[str, Any]]) -> Dict[str, List[Dict[str, Any]]]:
    branches = [(position, pillars[position]["zhi"]) for position in POSITIONS]
    out = {
        "liu_he": [],
        "half_he": [],
        "liu_chong": [],
        "hai_chuan": [],
        "po": [],
        "jue": [],
        "self_punish": [],
    }
    present_full_san_he = {key for key in BRANCH_SAN_HE if key <= {branch for _, branch in branches}}
    for left_index, (left_position, left_branch) in enumerate(branches):
        for right_position, right_branch in branches[left_index + 1 :]:
            key = frozenset((left_branch, right_branch))
            items = [
                _position_item(pillars, left_position, "支"),
                _position_item(pillars, right_position, "支"),
            ]
            if len(key) == 1 and left_branch in BRANCH_SELF_PUNISH:
                out["self_punish"].append(
                    _pair_relation("自刑", f"{left_branch}{right_branch}自刑", items, branches=[left_branch, right_branch])
                )
            if key in BRANCH_LIU_HE:
                name, transform, note = BRANCH_LIU_HE[key]
                out["liu_he"].append(
                    _pair_relation("六合", name, items, branches=[left_branch, right_branch], transform_element=transform, note=note)
                )
            if key in BRANCH_HALF_HE and not any(key < full_key for full_key in present_full_san_he):
                name, element, category = BRANCH_HALF_HE[key]
                out["half_he"].append(
                    _pair_relation("半合", name, items, branches=[left_branch, right_branch], element=element, category=category)
                )
            if key in BRANCH_CHONG:
                out["liu_chong"].append(
                    _pair_relation("六冲", BRANCH_CHONG[key], items, branches=[left_branch, right_branch])
                )
            if key in BRANCH_HAI_CHUAN:
                out["hai_chuan"].append(
                    _pair_relation("六害/穿", BRANCH_HAI_CHUAN[key], items, branches=[left_branch, right_branch])
                )
            if key in BRANCH_PO:
                out["po"].append(_pair_relation("破", BRANCH_PO[key], items, branches=[left_branch, right_branch]))
            if key in BRANCH_JUE:
                out["jue"].append(_pair_relation("绝", BRANCH_JUE[key], items, branches=[left_branch, right_branch]))
    return out


def _branch_group_relations(pillars: Dict[str, Dict[str, Any]]) -> Dict[str, List[Dict[str, Any]]]:
    present = {pillars[position]["zhi"] for position in POSITIONS}
    san_he = []
    san_hui = []
    three_punishments = []
    for key, (name, element) in BRANCH_SAN_HE.items():
        if key <= present:
            items = _branch_items_by_value(pillars, set(key))
            san_he.append(_pair_relation("三合", name, items, branches=sorted(key), element=element, complete=True))
    for key, (name, element) in BRANCH_SAN_HUI.items():
        if key <= present:
            items = _branch_items_by_value(pillars, set(key))
            san_hui.append(_pair_relation("三会", name, items, branches=sorted(key), element=element, complete=True))
    for key, name in BRANCH_THREE_PUNISHMENTS.items():
        present_part = key & present
        if len(present_part) >= 2:
            items = _branch_items_by_value(pillars, set(present_part))
            three_punishments.append(
                _pair_relation(
                    "三刑",
                    name if len(present_part) == 3 else name + "未全",
                    items,
                    branches=sorted(present_part),
                    complete=len(present_part) == 3,
                    missing=sorted(key - present_part),
                )
            )
    return {"san_he": san_he, "san_hui": san_hui, "three_punishments": three_punishments}


def _original_relations(pillars: Dict[str, Dict[str, Any]]) -> Dict[str, Any]:
    stem_relations = _stem_relations(pillars)
    branch_pairs = _branch_pair_relations(pillars)
    branch_groups = _branch_group_relations(pillars)
    branch_relations = {**branch_pairs, **branch_groups}
    return {
        "stem_relations": stem_relations,
        "branch_relations": branch_relations,
        "summary": {
            "stem_combinations": len(stem_relations["combinations"]),
            "stem_clashes": len(stem_relations["clashes"]),
            "branch_relation_count": sum(len(values) for values in branch_relations.values()),
        },
    }


def _day_master_element(day_master: str) -> str:
    for element, rule in ROOT_RULES.items():
        if day_master in rule["day_masters"]:
            return element
    return ""


def _season_bias(month_branch: str) -> Dict[str, str]:
    if month_branch in WINTER_BRANCHES:
        return {
            "type": "寒重",
            "useful_element": "火",
            "reason": "月令在亥子丑冬令，原局先看寒重，火为调候第一用神",
        }
    if month_branch in SUMMER_BRANCHES:
        return {
            "type": "燥重",
            "useful_element": "水",
            "reason": "月令在巳午未夏令，原局先看燥重，水为调候第一用神",
        }
    return {"type": "平", "useful_element": "", "reason": "月令不在寒燥极端季节，调候不优先于格局取用"}


def _ten_god_matrix(pillars: Dict[str, Dict[str, Any]]) -> Dict[str, Any]:
    stems = []
    branches = []
    counts: Dict[str, int] = {}
    for position in POSITIONS:
        pillar = pillars[position]
        stem_god = pillar["shi_shen_gan"]
        stems.append(
            {
                "position": position,
                "label": f"{POSITION_LABELS[position]}干",
                "gan": pillar["gan"],
                "shi_shen": stem_god,
            }
        )
        counts[stem_god] = counts.get(stem_god, 0) + 1
        hidden = []
        for gan, shi_shen in zip(pillar["hide_gan"], pillar["shi_shen_zhi"]):
            hidden.append({"gan": gan, "shi_shen": shi_shen})
            counts[shi_shen] = counts.get(shi_shen, 0) + 1
        branches.append(
            {
                "position": position,
                "label": f"{POSITION_LABELS[position]}支",
                "zhi": pillar["zhi"],
                "main_shi_shen": pillar["shi_shen_zhi"][0] if pillar["shi_shen_zhi"] else "",
                "hidden": hidden,
            }
        )
    return {"stems": stems, "branches": branches, "counts": counts}


def _root_summary(day_master: str, pillars: Dict[str, Dict[str, Any]]) -> Dict[str, Any]:
    element = _day_master_element(day_master)
    rule = ROOT_RULES.get(element, {})
    details = []
    found = []
    for position in POSITIONS:
        branch = pillars[position]["zhi"]
        types = []
        if branch in rule.get("bijie", set()):
            types.append("明现比劫")
        if branch in rule.get("yin", set()):
            types.append("明现印")
        if branch in rule.get("storehouse", set()):
            types.append("对应库")
        if branch in rule.get("support", set()):
            types.append("戊己土印星兼助力")
        if types:
            found.extend(types)
        details.append(
            {
                "position": position,
                "label": f"{POSITION_LABELS[position]}支",
                "zhi": branch,
                "root_types": types,
                "has_formal_root": bool(types),
            }
        )
    has_root = any(item["has_formal_root"] for item in details)
    if has_root:
        status = "有根"
        conclusion = "有根气，可担财官、用食伤"
    else:
        status = "无根"
        conclusion = "无正式根气，需看大运补根或从格条件"
    return {
        "day_master_element": element,
        "has_root": has_root,
        "status": status,
        "root_types": sorted(set(found)),
        "details": details,
        "conclusion": conclusion,
    }


def _tomb_markers(pillars: Dict[str, Dict[str, Any]]) -> List[Dict[str, Any]]:
    markers = []
    for position in POSITIONS:
        pillar = pillars[position]
        branch = pillar["zhi"]
        if branch in TOMB_BRANCHES:
            markers.append(
                {
                    "position": position,
                    "label": f"{POSITION_LABELS[position]}支",
                    "zhi": branch,
                    "hidden_gan": pillar["hide_gan"],
                    "shi_shen": pillar["shi_shen_zhi"],
                    "space": "家外/社会" if position in {"year", "month"} else "家里/自己",
                }
            )
    return markers


def _qisha_marker(pillars: Dict[str, Dict[str, Any]]) -> Dict[str, Any]:
    occurrences = []
    has_qisha_stem = False
    has_qisha_main_root = False
    for position in POSITIONS:
        pillar = pillars[position]
        if pillar["shi_shen_gan"] == "七杀":
            has_qisha_stem = True
            occurrences.append({"position": position, "label": f"{POSITION_LABELS[position]}干", "value": pillar["gan"]})
        if pillar["shi_shen_zhi"] and pillar["shi_shen_zhi"][0] == "七杀":
            has_qisha_main_root = True
        for gan, shi_shen in zip(pillar["hide_gan"], pillar["shi_shen_zhi"]):
            if shi_shen == "七杀":
                occurrences.append({"position": position, "label": f"{POSITION_LABELS[position]}支藏干", "value": gan})
    month_main_qisha = bool(pillars["month"]["shi_shen_zhi"] and pillars["month"]["shi_shen_zhi"][0] == "七杀")
    strong = month_main_qisha or (has_qisha_stem and has_qisha_main_root)
    return {
        "exists": bool(occurrences),
        "strong": strong,
        "occurrences": occurrences,
        "priority_note": "月令本气为七杀，或七杀透出且地支有七杀本气根，分析时有杀先论杀" if strong else "",
    }


def _analysis_prep(pillars: Dict[str, Dict[str, Any]], day_master: str) -> Dict[str, Any]:
    month = pillars["month"]
    month_branch = month["zhi"]
    month_ling_shishen = month["shi_shen_zhi"][0] if month["shi_shen_zhi"] else ""
    tombs = _tomb_markers(pillars)
    qisha = _qisha_marker(pillars)
    season = _season_bias(month_branch)
    priority_scan = []
    if tombs:
        priority_scan.append("有四墓库，先检查是否形成库做功")
    if qisha["strong"]:
        priority_scan.append("七杀较强，先按有杀先论杀处理")
    if season["type"] in {"寒重", "燥重"}:
        priority_scan.append(f"原局{season['type']}，先看{season['useful_element']}调候")
    if not priority_scan:
        priority_scan.append("无明显高优先级特殊分支，按月令十神定格局取用")
    return {
        "day_master": day_master,
        "month_ling": month_branch,
        "month_ling_shishen": month_ling_shishen,
        "month_ling_note": "月令本气十神为原局最强十神，是普通格局取用的主轴",
        "ten_god_matrix": _ten_god_matrix(pillars),
        "root_summary": _root_summary(day_master, pillars),
        "season_bias": season,
        "tomb_markers": tombs,
        "qisha_marker": qisha,
        "original_relations": _original_relations(pillars),
        "priority_scan": priority_scan,
    }


def generate_chart(input_data: Dict[str, Any] | BaziRequest) -> Dict[str, Any]:
    req = input_data if isinstance(input_data, BaziRequest) else request_from_dict(input_data)
    solar, lunar = _solar_lunar_from_request(req)
    eight_char = lunar.getEightChar()
    eight_char.setSect(req.sect)
    yun = eight_char.getYun(gender_number(req.gender), req.yun_sect)

    pillars = {
        "year": _pillar(eight_char, "year"),
        "month": _pillar(eight_char, "month"),
        "day": _pillar(eight_char, "day"),
        "time": _pillar(eight_char, "time"),
    }
    day_master = eight_char.getDayGan()
    da_yun = []
    for dy in yun.getDaYun(req.da_yun_count + 1):
        da_yun.append(
            {
                "index": dy.getIndex(),
                "name": "起运前" if dy.getIndex() == 0 else f"第{dy.getIndex()}步大运",
                "gan_zhi": dy.getGanZhi(),
                "start_year": dy.getStartYear(),
                "end_year": dy.getEndYear(),
                "start_age": dy.getStartAge(),
                "end_age": dy.getEndAge(),
                "xun": dy.getXun() if dy.getIndex() > 0 else "",
                "xun_kong": dy.getXunKong() if dy.getIndex() > 0 else "",
                "liu_nian": _liu_nian_list(dy),
            }
        )

    result = {
        "name": req.name,
        "gender": "男" if req.gender == "male" else "女",
        "calendar_type": "阳历" if req.calendar_type == "solar" else "农历",
        "input": req.__dict__,
        "solar": {
            "text": solar.toYmdHms(),
            "year": solar.getYear(),
            "month": solar.getMonth(),
            "day": solar.getDay(),
            "hour": solar.getHour(),
            "minute": solar.getMinute(),
            "second": solar.getSecond(),
        },
        "lunar": {
            "text": _fmt_lunar(lunar),
            "full_text": lunar.toFullString(),
            "year": lunar.getYear(),
            "month": lunar.getMonth(),
            "day": lunar.getDay(),
            "is_leap_month": lunar.getMonth() < 0,
        },
        "bazi": {
            "text": eight_char.toString(),
            "year": pillars["year"]["gan_zhi"],
            "month": pillars["month"]["gan_zhi"],
            "day": pillars["day"]["gan_zhi"],
            "time": pillars["time"]["gan_zhi"],
            "day_master": day_master,
            "pillars": pillars,
            "tai_yuan": eight_char.getTaiYuan(),
            "tai_xi": eight_char.getTaiXi(),
            "ming_gong": eight_char.getMingGong(),
            "shen_gong": eight_char.getShenGong(),
        },
        "analysis_prep": _analysis_prep(pillars, day_master),
        "yun": {
            "is_forward": yun.isForward(),
            "direction": "顺排" if yun.isForward() else "逆排",
            "start_text": _fmt_start(yun),
            "start_solar": yun.getStartSolar().toYmdHms(),
            "da_yun": da_yun,
        },
        "rules": {
            "sect": req.sect,
            "sect_desc": "晚子时日柱算当天" if req.sect == 2 else "晚子时日柱算明天",
            "yun_sect": req.yun_sect,
            "yun_sect_desc": "按天数和时辰数起运" if req.yun_sect == 1 else "按分钟数起运",
            "time_standard": "北京时间/输入时间；未做真太阳时校正",
        },
    }
    result["plain_text"] = chart_to_plain_text(result)
    return result


def chart_to_plain_text(chart: Dict[str, Any]) -> str:
    bazi = chart["bazi"]
    yun = chart["yun"]
    prep = chart.get("analysis_prep", {})
    root = prep.get("root_summary", {})
    season = prep.get("season_bias", {})
    relation_summary = (prep.get("original_relations") or {}).get("summary", {})
    lines = []
    if chart.get("name"):
        lines.append(f"姓名：{chart['name']}")
    lines.extend(
        [
            f"性别：{chart['gender']}",
            f"阳历：{chart['solar']['text']}",
            f"农历：{chart['lunar']['text']}",
            "",
            "四柱：",
            f"年柱：{bazi['year']}",
            f"月柱：{bazi['month']}",
            f"日柱：{bazi['day']}（日主：{bazi['day_master']}）",
            f"时柱：{bazi['time']}",
            "",
            "程序预分析：",
            f"月令：{prep.get('month_ling', '')}（本气十神：{prep.get('month_ling_shishen', '')}）",
            f"根气：{root.get('conclusion', '')}",
            f"寒燥：{season.get('type', '')} {season.get('useful_element', '')}".rstrip(),
            "原局关系："
            f"天干合{relation_summary.get('stem_combinations', 0)}组、"
            f"天干冲{relation_summary.get('stem_clashes', 0)}组、"
            f"地支关系{relation_summary.get('branch_relation_count', 0)}组",
            "特殊优先级：" + "；".join(prep.get("priority_scan", [])),
            "",
            f"起运：{yun['start_text']}，{yun['start_solar']} 起运，{yun['direction']}",
            "大运：",
        ]
    )
    for dy in yun["da_yun"]:
        if dy["index"] == 0:
            continue
        lines.append(
            f"{dy['start_age']:02d}-{dy['end_age']:02d}岁 "
            f"{dy['start_year']}-{dy['end_year']}：{dy['gan_zhi']}"
        )
    return "\n".join(lines)


def read_batch_rows(file_name: str, data: bytes) -> List[Dict[str, Any]]:
    lower = file_name.lower()
    if lower.endswith(".csv"):
        text = data.decode("utf-8-sig")
        return [dict(row) for row in csv.DictReader(StringIO(text))]
    if lower.endswith(".xlsx"):
        try:
            from openpyxl import load_workbook
        except ImportError as exc:
            raise RuntimeError("当前 Python 环境缺少 openpyxl，无法读取 xlsx") from exc
        wb = load_workbook(BytesIO(data), data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return []
        headers = [str(v or "").strip() for v in rows[0]]
        output = []
        for row in rows[1:]:
            output.append({headers[i]: row[i] if i < len(row) else "" for i in range(len(headers))})
        return output
    raise ValueError("批量导入只支持 .csv 或 .xlsx")


def batch_generate(rows: List[Dict[str, Any]], defaults: Optional[Dict[str, Any]] = None) -> List[Dict[str, Any]]:
    defaults = defaults or {}
    output = []
    for index, row in enumerate(rows, start=1):
        merged = dict(defaults)
        merged.update({k: v for k, v in row.items() if v is not None and str(v).strip() != ""})
        try:
            chart = generate_chart(merged)
            output.append({"row": index, "ok": True, "chart": chart, "error": ""})
        except Exception as exc:  # noqa: BLE001
            output.append({"row": index, "ok": False, "chart": None, "error": str(exc)})
    return output


def export_batch_xlsx(results: List[Dict[str, Any]]) -> bytes:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font
    except ImportError as exc:
        raise RuntimeError("当前 Python 环境缺少 openpyxl，无法导出 xlsx") from exc

    wb = Workbook()
    ws = wb.active
    ws.title = "八字排盘"
    headers = [
        "行号",
        "状态",
        "错误信息",
        "姓名",
        "性别",
        "阳历",
        "农历",
        "年柱",
        "月柱",
        "日柱",
        "时柱",
        "日主",
        "起运",
        "起运阳历",
        "顺逆",
        "大运简表",
        "排盘文本",
    ]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")

    for item in results:
        if not item["ok"]:
            ws.append([item["row"], "失败", item["error"]])
            continue
        chart = item["chart"]
        bazi = chart["bazi"]
        yun = chart["yun"]
        da_yun_text = "；".join(
            f"{dy['start_age']}-{dy['end_age']}岁 {dy['start_year']}-{dy['end_year']} {dy['gan_zhi']}"
            for dy in yun["da_yun"]
            if dy["index"] > 0
        )
        ws.append(
            [
                item["row"],
                "成功",
                "",
                chart.get("name", ""),
                chart["gender"],
                chart["solar"]["text"],
                chart["lunar"]["text"],
                bazi["year"],
                bazi["month"],
                bazi["day"],
                bazi["time"],
                bazi["day_master"],
                yun["start_text"],
                yun["start_solar"],
                yun["direction"],
                da_yun_text,
                chart["plain_text"],
            ]
        )

    widths = [8, 8, 24, 12, 8, 22, 28, 8, 8, 8, 8, 8, 18, 22, 8, 60, 60]
    for i, width in enumerate(widths, start=1):
        ws.column_dimensions[chr(64 + i)].width = width
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def json_dumps(data: Dict[str, Any]) -> str:
    return json.dumps(data, ensure_ascii=False, indent=2)
